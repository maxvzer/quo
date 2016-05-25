# coding: utf-8
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.views.generic import TemplateView
from django.core.urlresolvers import reverse
from models import Person, MilitaryCommissariat,  Grade,\
                   DegreeOfFitness, Faculty, Department, Group, OKSO, VUS, Year

from django.core.exceptions import ObjectDoesNotExist
import openpyxl
from openpyxl.styles import Border, Font, Alignment, Side
import sys
import os
from django import forms
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.db import connection
from django.db.models import Q

# Create your views here.
from Quo.constants import DEPARTMENT_TYPES, MILITARY_COMISSARIAT_TYPES, GRADE_TYPES, VUS_TYPES


class IndexView(TemplateView):
    template_name = "main/index.html"
    path = 'D:\Programs\Python\Projects\quo\LOLSave.xlsx'

    def post(self, request, *args, **kwargs):
        if self.request.method == 'POST':
            if 'big_list' in request.POST:
                students = Person.objects.all()
                result = []
                for student in students:
                    _student = {}
                    _student.update({'first_name': student.first_name,
                                     'last_name': student.last_name,
                                     'middle_name': student.middle_name,
                                     'VUS': student.VUS.VUS,
                                     'group': student.group.group,
                                     'faculty': student.faculty.faculty,
                                     'department': student.department.department,
                                     'birth_date': str(student.birth_date),
                                     'OKSO': student.OKSO.OKSO,
                                     'address': student.address,
                                     'title': student.military.title,
                                     'birth_place': student.birthplace,
                                     'document_number' : student.document_number,
                                     'issued_by': student.issued_by,
                                     'degree_of_fitness': student.degree.degree_of_fitness,
                                     'grade': student.grade.grade,
                                     'year': student.year.year})
                    result.append(_student)
                print result
                big_list(result,self.path)
                return HttpResponseRedirect('/')
            elif 'abiturients' in request.POST:
                students = Person.objects.filter(is_learning=False)
                result = []
                for student in students:
                    _student = {}
                    _student.update({'first_name': student.first_name,
                                     'last_name': student.last_name,
                                     'middle_name': student.middle_name,
                                     'VUS': student.VUS.VUS,
                                     'group': student.group.group,
                                     'faculty': student.faculty.faculty,
                                     'department': student.department.department,
                                     'birth_date': str(student.birth_date),
                                     'OKSO': student.OKSO.OKSO,
                                     'address': student.address,
                                     'title': student.military.title,
                                     'birth_place': student.birthplace,
                                     'document_number' : student.document_number,
                                     'issued_by': student.issued_by,
                                     'degree_of_fitness': student.degree.degree_of_fitness,
                                     'grade': student.grade.grade,
                                     'year': student.year.year})
                    result.append(_student)
                print result
                table_submission_of_an_application(result,self.path)
                return HttpResponseRedirect('/')
            elif 'VUS' in request.POST:
                students = Person.objects.filter(is_learning=True)
                result = []
                for student in students:
                    _student = {}
                    _student.update({'first_name': student.first_name,
                                     'last_name': student.last_name,
                                     'middle_name': student.middle_name,
                                     'VUS': student.VUS.VUS,
                                     'group': student.group.group,
                                     'faculty': student.faculty.faculty,
                                     'department': student.department.department,
                                     'birth_date': str(student.birth_date),
                                     'OKSO': student.OKSO.OKSO,
                                     'address': student.address,
                                     'title': student.military.title,
                                     'birth_place': student.birthplace,
                                     'document_number' : student.document_number,
                                     'issued_by': student.issued_by,
                                     'degree_of_fitness': student.degree.degree_of_fitness,
                                     'grade': student.grade.grade,
                                     'year': int(student.year.year)})
                    result.append(_student)
                print result
                list_for_vus(result,self.path)
                return HttpResponseRedirect('/')

            elif 'training' in request.POST:
                students = Person.objects.filter(VUS__VUS='084000', is_learning=True)
                result = []
                for student in students:
                    _student = {}
                    _student.update({'first_name': student.first_name,
                                     'last_name': student.last_name,
                                     'middle_name': student.middle_name,
                                     'VUS': student.VUS.VUS,
                                     'group': student.group.group,
                                     'faculty': student.faculty.faculty,
                                     'department': student.department.department,
                                     'birth_date': str(student.birth_date),
                                     'OKSO': student.OKSO.OKSO,
                                     'address': student.address,
                                     'title': student.military.title,
                                     'birth_place': student.birthplace,
                                     'document_number': student.document_number,
                                     'issued_by': student.issued_by,
                                     'degree_of_fitness': student.degree.degree_of_fitness,
                                     'grade': student.grade.grade,
                                     'year': int(student.year.year)})
                    result.append(_student)
                print result
                table_military_training(result, self.path)
                return HttpResponseRedirect('/')

            elif 'OVK' in request.POST:
                return HttpResponseRedirect('/')


class ColumnsView(TemplateView):
    template_name = "main/columns.html"

    def post(self, request, *args, **kwargs):
        if self.request.method == 'POST':
                some_var = self.request.POST.getlist('checks[]')
                print some_var
                url = "../filters/?columns=" + str(some_var)
                return HttpResponseRedirect(url)


class StudentView(TemplateView):
    template_name = "main/students.html"

    def post(self, request, *args, **kwargs):
        if self.request.method == 'POST':
            students = self.request.POST.getlist('students[]')
            for student in students:
                year = Year.objects.get(year=1)
                Person.objects.filter(id=student).update(is_learning=True, year=year)
            url = "../"
            return HttpResponseRedirect(url)

    def get_context_data(self, **kwargs):
        context = super(StudentView, self).get_context_data(**kwargs)
        students = Person.objects.filter(is_learning=False)
        context.update({
                'students': students
            })
        return context


class ErrorsView(TemplateView):
    template_name = "main/errors.html"

    def get_context_data(self, **kwargs):
        context = super(ErrorsView, self).get_context_data(**kwargs)
        students = Person.objects.filter(Q(military__title="none") | Q(grade__grade="none"))
        context.update({
                'students': students
            })
        return context


class FiltersView(TemplateView):
    template_name = "main/filters.html"

    def get_context_data(self, **kwargs):
        context = super(FiltersView, self).get_context_data(**kwargs)
        columns = self.request.GET.get('columns').translate({ord(i): None for i in "[]u' "})
        columns_array = []
        for column in columns.split(','):
            columns_array.append(column)
        print columns_array[0]

        persons = Person.objects.all()
        militarys = MilitaryCommissariat.objects.all()
        grades = Grade.objects.all()
        degrees = DegreeOfFitness.objects.all()
        facultys = Faculty.objects.all()
        departments = Department.objects.all()
        groups = Group.objects.all()
        oksos = OKSO.objects.all()
        VUSes = VUS.objects.all()
        years = Year.objects.all()

        context.update({
            'columns': columns_array,
            'persons': persons,
            'militarys': militarys,
            'grades': grades,
            'degrees': degrees,
            'facultys': facultys,
            'departments': departments,
            'groups': groups,
            'oksos': oksos,
            'VUSes': VUSes,
            'years': years
        })
        return context

    def post(self, request, *args, **kwargs):
        if self.request.method == 'POST':

            columns = self.request.GET.get('columns')
            group = self.request.POST['group']
            department = self.request.POST['department']
            faculty = self.request.POST['faculty']
            okso = self.request.POST['okso']
            military = self.request.POST['military']
            degree = self.request.POST['degree']
            grade = self.request.POST['grade']
            VUS = self.request.POST['VUS']
            year = self.request.POST['year']
            filters = []
            filters.extend([faculty,department,group,okso,military,degree,grade,VUS,year])

            # filters = "" + faculty + department + group + okso + military + degree + grade

            url = "../result/?columns=" + str(columns) + "&filters=" + str(filters)

            # for var in some_var:
            #     url = url + u"value" + str(number) + u"=" + var + u"&"
            #     number += 1
            return HttpResponseRedirect(url)


class ResultView(TemplateView):
    template_name = "main/result.html"

    def dictfetchall(self, cursor):
        "Return all rows from a cursor as a dict"
        columns = [col[0] for col in cursor.description]
        return [
            dict(zip(columns, row))
            for row in cursor.fetchall()
        ]

    def get_context_data(self, **kwargs):
        context = super(ResultView, self).get_context_data(**kwargs)

        query = '''SELECT '''

        columns_dict = {
            'FIO': 'main_person.first_name, main_person.middle_name, main_person.last_name,',
            'birth_date': 'main_person.birth_date,',
            'grop': 'main_group.group, main_faculty.faculty, main_department.department,',
            'grade': 'main_grade.grade,',
            'military': 'main_militarycommissariat.title,',
            'fitness': 'main_degreeoffitness.degree_of_fitness,',
            'passport': 'main_person.document_number, main_person.issued_by,',
            'OKSO': 'main_okso.OKSO,',
            'VUS': 'main_vus.VUS,',
            'year': 'main_year.year,',
            'birthplace': 'main_person.birthplace,',
            'address': 'main_person.address,',
            'OKSO_letters': ''
        }

        filters_dict = {
            '2': 'main_group.group = %s and ',
            '1': 'main_department.department = %s and ',
            '0': 'main_faculty.faculty = %s and ',
            '4': 'main_militarycommissariat.title = %s and ',
            '3': 'main_okso.OKSO = %s and ',
            '6': 'main_grade.grade = %s and ',
            '5': 'main_degreeoffitness.degree_of_fitness = %s and ',
            '7': 'main_vus.VUS = %s and ',
            '8': 'main_year.year = %s and ',
            'None': ''

        }

        columns = self.request.GET.get('columns').translate({ord(i): None for i in "[]u' "})
        columns_array = []
        for column in columns.split(','):
            columns_array.append(column)
            try:
                query_part = columns_dict[column]
            except KeyError as e:

                raise ValueError('Undefined unit: {}'.format(e.args[0]))
            query += query_part
        query = query[:-1] + " "

        query += '''FROM main_person
                    INNER JOIN main_militarycommissariat
                    ON main_person.military_id = main_militarycommissariat.id
                    INNER JOIN main_degreeoffitness
                    ON main_person.degree_id = main_degreeoffitness.id
                    INNER JOIN main_okso
                    ON main_person.okso_id = main_okso.id
                    INNER JOIN main_grade
                    ON main_person.grade_id = main_grade.id
                    INNER JOIN main_group
                    ON main_person.group_id = main_group.id
                    INNER JOIN main_department
                    ON main_person.department_id = main_department.id
                    INNER JOIN main_faculty
                    ON main_person.faculty_id = main_faculty.id
                    INNER JOIN main_vus
                    ON main_person.VUS_id = main_vus.id
                    INNER JOIN main_year
                    ON main_person.year_id = main_year.id
                    '''

        if self.request.GET.get('filters'):
            query += '''WHERE '''

        filters = self.request.GET.get('filters').translate({ord(i): None for i in "[]u' "})
        filters_array = []
        number = 0
        params = []
        for _filter in filters.split(','):
            if _filter == "None":
                number += 1
                continue
            filters_array.append(_filter)
            try:
                query_part = filters_dict[str(number)]
                params.append(_filter)
            except KeyError as e:
                raise ValueError('Undefined unit: {}'.format(e.args[0]))
            query += query_part
            number += 1
        query = query[:-4] + " "

        print query + str(params) + "\n\n"

        with connection.cursor() as cursor:
            cursor.execute(query, params)
            values = self.dictfetchall(cursor)
            print values

        facultys_dict = dict(DEPARTMENT_TYPES)  # Conversion to a dictionary mapping
        militarys_dict = dict(MILITARY_COMISSARIAT_TYPES)
        grades_dict = dict(GRADE_TYPES)
        for dicti in values:
            try:
                dicti['faculty'] = facultys_dict[dicti.get('faculty')]
            except:
                pass
            try:
                dicti['title'] = militarys_dict[dicti.get('title')]
            except:
                pass
            try:
                dicti['grade'] = grades_dict[dicti.get('grade')]
            except:
                pass

        context.update({
            'values': values,
            'columns': columns_array,
            'length': len(values)
        })
        return context


class UploadFileForm(forms.Form):
    file = forms.FileField()


def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            handle_uploaded_file(request.FILES['file'])
            return HttpResponseRedirect('/')
    else:
        form = UploadFileForm()
    return render_to_response('main/upload_form.html', {'form': form}, RequestContext(request))


def handle_uploaded_file(f):
    with open('test.xlsx', 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)

    parse_table('test.xlsx', u'Лист1')
    # wb = openpyxl.load_workbook(filename='test.xlsx')
    # sheet = wb['Worksheet']
    # print >> sys.stderr, sheet['A1'].value
    # # os.remove('D:\\Programs\\Python\\Projects\\MyFirst\\temp.xls')
    # print >> sys.stderr, u'Я смог'


# Функция для получения следующего символа ячейки
def move(letter):
    if letter.lower() == 'z':
        return unichr(97)
    return unichr(ord(letter.lower())+1)


# функция получени столбца по заданному имени
def check_print(my_field, sheet, target_list):

    a = 1
    b = 'A'
    my_field_temp = my_field.replace(' ', '')
    val = sheet[b + str(a)].value
    while val is not None:
        b = move(b).upper()
        val = sheet[b+str(a)].value
        if(val is not None):
            val_temp = val.replace(' ', '')
        else:
            break
        if val_temp.upper() == my_field_temp.upper():     # Исправил, чтобы по верхнему регистру проверял
            a = 2
            val = sheet[b+str(a)].value
            while val is not None:
                target_list.append(val)
                a += 1
                m = b + str(a)
                val = sheet[m].value


def parse_table(filename_in, sheet_in):

    wb_funk = openpyxl.load_workbook(filename=filename_in)
    sheet = wb_funk[sheet_in]
    fio = []
    fio_rus = u'Фамилия Имя Отчество'
    group_all = []
    group_rus = u'Учебная группа'
    special = []
    special_rus = u'Специальность'
    birthday = []
    birthday_rus = u'Дата рождения'
    military_office = []
    military_office_rus = u'Военкомат'
    home_address = []
    home_address_rus = u'Домашний адрес (по паспорту)'
    place_of_birth = []
    place_of_birth_rus = u'Место рождения'
    passport_number = []
    passport_number_rus = u'Серия и номер паспорта'
    issued = []
    issued_rus = u'Кем и когда выдан паспорт'
    the_degree_of_validity = []
    the_degree_of_validity_rus = u'Степень годности'
    list_vus = []
    vus_rus = u'ВУС'
    rank = []

    list_vus_officer = ['084000', '085000', '141600']
    list_vus_sage = ['659182']
    list_vus_soldier = ['659995', '059995']

    check_print(fio_rus, sheet, fio)
    check_print(group_rus, sheet, group_all)
    check_print(special_rus, sheet, special)
    check_print(birthday_rus, sheet, birthday)
    check_print(military_office_rus, sheet, military_office)
    check_print(home_address_rus, sheet, home_address)
    check_print(place_of_birth_rus, sheet, place_of_birth)
    check_print(passport_number_rus, sheet, passport_number)
    check_print(issued_rus, sheet, issued)
    check_print(the_degree_of_validity_rus, sheet, the_degree_of_validity)
    check_print(vus_rus, sheet, list_vus)

    group = []
    department = []
    faculty = []
    for gr in group_all:
        str_number = ''
        str_faculty_temp = ''
        flag1 = False
        for i in gr:
            if (i >= '0') and (i <= '9'):
                if flag1 is False:
                    faculty.append(str_faculty_temp)
                    flag1 = True
                    str_number += i
                else:
                    str_number += i
            else:
                str_faculty_temp += i
            if i == '-':
                department.append(str_number)
                str_number = ''
        group.append(str_number)

    firstname = []
    middlename = []
    lastname = []
    for sub_fio in fio:
        list_temp = sub_fio.split(' ')
        lastname.append(list_temp[0])
        firstname.append(list_temp[1])
        middlename.append(list_temp[2])

    for vus in list_vus:
        if vus in list_vus_officer:
            rank.append(u'Офицер')
        elif vus in list_vus_soldier:
            rank.append(u'Солдат')
        else:
            rank.append(u'Сержант')

    number_of_record = len(firstname)
    i = 0
    while i < number_of_record:
        try:
            okso = OKSO.objects.get(OKSO=special[i])
        except ObjectDoesNotExist:
            okso = OKSO.objects.create(OKSO=special[i])

        try:
            _group = Group.objects.get(group=group[i])
        except ObjectDoesNotExist:
            _group = Group.objects.create(group=group[i])
        try:
            _department = Department.objects.get(department=department[i])
        except ObjectDoesNotExist:
            _department = Department.objects.create(department=department[i])

        faculty_eng = "none"
        for value in DEPARTMENT_TYPES:
            if faculty[i] == value[1]:
                faculty_eng = value[0]

        try:
            _faculty = Faculty.objects.get(faculty=faculty_eng)
        except ObjectDoesNotExist:
            _faculty = Faculty.objects.create(faculty=faculty_eng)

        try:
            vus = VUS.objects.get(VUS=list_vus[i])
        except ObjectDoesNotExist:
            vus = VUS.objects.create(VUS=list_vus[i])

        try:
            degree = DegreeOfFitness.objects.get(degree_of_fitness=the_degree_of_validity[i])
        except ObjectDoesNotExist:
            degree = DegreeOfFitness.objects.create(degree_of_fitness=the_degree_of_validity[i])

        try:
            year = Year.objects.get(year=0)
        except ObjectDoesNotExist:
            year = Year.objects.create(year=0)

        militarys = MILITARY_COMISSARIAT_TYPES

        for _tuple in militarys:
            if _tuple[1] == military_office[i] or military_office[i] in _tuple[1]:
                try:
                    military = MilitaryCommissariat.objects.get(title=_tuple[0])
                except ObjectDoesNotExist:
                    military = MilitaryCommissariat.objects.create(title=_tuple[0])
            else:
                try:
                    military = MilitaryCommissariat.objects.get(title=u'none')
                except ObjectDoesNotExist:
                    military = MilitaryCommissariat.objects.create(title=u'none')

        if rank[i] == u'Сержант':
            try:
                _rank = Grade.objects.get(grade='sergeant')
            except ObjectDoesNotExist:
                _rank = Grade.objects.create(grade='sergeant')
        elif rank[i] == u'Офицер':
            try:
                _rank = Grade.objects.get(grade='officer')
            except ObjectDoesNotExist:
                _rank = Grade.objects.create(grade='officer')
        elif rank[i] == u'Солдат':
            try:
                _rank = Grade.objects.get(grade='soldier')
            except ObjectDoesNotExist:
                _rank = Grade.objects.create(grade='soldier')
        else:
            try:
                _rank = Grade.objects.get(grade='none')
            except ObjectDoesNotExist:
                _rank = Grade.objects.create(grade='none')

        person = Person.objects.create(first_name=firstname[i],
                                       middle_name=middlename[i],
                                       last_name=lastname[i],
                                       birth_date=birthday[i],
                                       OKSO=okso,
                                       faculty=_faculty,
                                       department=_department,
                                       group=_group,
                                       address=home_address[i],
                                       birthplace=place_of_birth[i],
                                       issued_by=issued[i],
                                       VUS=vus,
                                       degree=degree,
                                       document_number=passport_number[i],
                                       year=year,
                                       # rank=rank,
                                       military=military,
                                       grade=_rank
                                       )


        # print (str(i+1)),
        # print firstname[i],
        # print middlename[i],
        # print lastname[i],
        # print special[i],
        # print birthday[i],
        # print faculty[i],
        # print department[i],
        # print group[i],
        # print home_address[i],
        # print place_of_birth[i],
        # print issued[i],
        # print list_vus[i],
        # print the_degree_of_validity[i],
        # print passport_number[i],
        # print military_office[i],
        # print rank[i]
        i += 1



# parse_table(lol, lol_sheet)

def find_counter_in_list(name, list_of_dict):
    n = len(list_of_dict)
    count = 0
    while count < n:
        sleeve = list_of_dict[count]['last_name'] + ' ' + list_of_dict[count]['first_name'] + ' ' +\
                 list_of_dict[count]['middle_name']
        if name == sleeve:
            return count
        count += 1

def print_table_header_vus(ws, cell_pos):

    font_zirn = Font(name='Times New Roman',
                     size=14,
                     bold=True,  # жирный
                     italic=False,  # курсив
                     vertAlign='baseline',  # по какому краю выравнивать
                     # (['subscript', 'baseline', 'superscript'])
                     underline='none',  # подчеркивание написанного
                     strike=False,  # зачеркнуть
                     color='FF000000')

    # выравнивание для шапки таблицы
    alignment_centre = Alignment(horizontal='center',  # выравнивание по горизонтали
                                 vertical='center',  # выравнивание по вертикали
                                 text_rotation=0,
                                 wrap_text=False,  # перенос текста
                                 shrink_to_fit=False,
                                 indent=0)


    # границы заливки (заливка таблицы)
    border1 = Border(left=Side(border_style='thin',
                               color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000'))

    cells_position = 'A' + str(cell_pos)
    cell_prop = ws[cells_position]
    cell_prop.value = u'№ п/п'
    cell_prop.font = font_zirn
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cells_position = 'B' + str(cell_pos)
    cell_prop = ws[cells_position]
    cell_prop.value = u'ФИО'
    cell_prop.font = font_zirn
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cells_position = 'C' + str(cell_pos)
    cell_prop = ws[cells_position]
    cell_prop.value = u'Группа'
    cell_prop.font = font_zirn
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_pos += 1

    return cell_pos

def table_submission_of_an_application(list_of_dict,path):

    lol_save = path
    lol_sheet = 'Worksheet'
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = lol_sheet

    # Шрифты для списка студентов(самый верх) и шапки колонки (Жирный)
    font2 = Font(name='Times New Roman',
                 size=14,
                 bold=True,  # жирный
                 italic=False,  # курсив
                 vertAlign='baseline',  # по какому краю выравнивать
                 # (['subscript', 'baseline', 'superscript'])
                 underline='none',  # подчеркивание написанного
                 strike=False,  # зачеркнуть
                 color='FF000000')

    # шрифты для подавших заявление на ВК2..... (не жирный)
    font3 = Font(name='Times New Roman',
                 size=14,
                 bold=False,  # жирный
                 italic=False,  # курсив
                 vertAlign='baseline',  # по какому краю выравнивать
                 # (['subscript', 'baseline', 'superscript'])
                 underline='none',  # подчеркивание написанного
                 strike=False,  # зачеркнуть
                 color='FF000000')

    # выравнивание для шапки таблицы
    alignment1 = Alignment(horizontal='center',  # выравнивание по горизонтали
                           vertical='center',  # выравнивание по вертикали
                           text_rotation=0,
                           wrap_text=False,  # перенос текста
                           shrink_to_fit=False,
                           indent=0)

    # выравнивание (ыравние по центру)
    alignment2 = Alignment(horizontal='center',  # выравнивание по горизонтали
                           vertical='center',  # выравнивание по вертикали
                           text_rotation=0,
                           wrap_text=False,  # перенос текста
                           shrink_to_fit=False,
                           indent=0)

    # выравнивание (ыравние по центру)
    alignment3 = Alignment(horizontal='left',  # выравнивание по горизонтали
                           vertical='center',  # выравнивание по вертикали
                           text_rotation=0,
                           wrap_text=False,  # перенос текста
                           shrink_to_fit=False,
                           indent=0)

    # границы заливки (заливка таблицы)
    border1 = Border(left=Side(border_style='thin',
                               color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000'))

    # Ширина колонок
    _cell = ws1.cell(row=1, column=1)
    _cell = ws1.cell(row=1, column=2)
    _cell = ws1.cell(row=1, column=3)
    _cell = ws1.cell(row=1, column=4)

    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 48
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 16

    # Заполнение клонок
    ws1.merge_cells('A1:D1')
    cell_prop = ws1['A1']
    cell_prop.value = u'Список студентов'
    cell_prop.font = font2
    cell_prop.alignment = alignment1

    ws1.merge_cells('A2:D2')
    cell_prop = ws1['A2']
    cell_prop.value = u'подавших заявления на ВК №2 для участия в конкурсном отборе'
    cell_prop.font = font3
    cell_prop.alignment = alignment1

    ws1.merge_cells('A3:D3')
    cell_prop = ws1['A3']
    cell_prop.value = u'на программы подготовки офицеров запаса'
    cell_prop.font = font3
    cell_prop.alignment = alignment1

    cell_prop = ws1['A5']
    cell_prop.value = u'№'
    cell_prop.font = font2
    cell_prop.alignment = alignment1
    cell_prop.border = border1

    cell_prop = ws1['B5']
    cell_prop.value = u'Фамилия, имя, отчество'
    cell_prop.font = font2
    cell_prop.alignment = alignment1
    cell_prop.border = border1

    cell_prop = ws1['C5']
    cell_prop.value = u'Группа'
    cell_prop.font = font2
    cell_prop.alignment = alignment1
    cell_prop.border = border1

    cell_prop = ws1['D5']
    cell_prop.value = u'Примечание'
    cell_prop.font = font2
    cell_prop.alignment = alignment1
    cell_prop.border = border1

    # получение списка групп и сортировка его
    s = []
    count = 0
    n = len(list_of_dict)

    list_full_group = []
    while count < n:
        my_gr = list_of_dict[count]['faculty'] + list_of_dict[count]['department'] +\
        '-' + list_of_dict[count]['group']
        list_full_group.append(my_gr)
        count += 1

    count = 0
    while count < n:
        if not (list_full_group[count] in s):
            s.append(list_full_group[count])
        count += 1
    s.sort()

    # вывод таблицы
    count_for_print = 6
    for group in s:
        count = 0
        first_last_three_names = []
        while count < n:
            if group == list_full_group[count]:
                sleeve = list_of_dict[count]['last_name'] + ' ' + list_of_dict[count]['first_name'] + ' ' +\
                         list_of_dict[count]['middle_name']
                first_last_three_names.append(sleeve)
            count += 1
        first_last_three_names.sort()
        for name in first_last_three_names:

            cell_cell = 'A' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = count_for_print - 5
            cell_prop.font = font3
            cell_prop.alignment = alignment1
            cell_prop.border = border1

            cell_cell = 'B' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = name
            cell_prop.font = font3
            cell_prop.alignment = alignment3
            cell_prop.border = border1

            cell_cell = 'C' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = group
            cell_prop.font = font3
            cell_prop.alignment = alignment2
            cell_prop.border = border1

            cell_cell = 'D' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.font = font3
            cell_prop.border = border1

            count_for_print += 1

    # Конец
    count_for_print += 1
    cell_cell = 'A' + str(count_for_print) + ':' + 'D' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'Начальник военной кафедры №2'
    cell_prop.font = font3
    cell_prop.alignment = alignment1

    count_for_print += 1
    cell_cell = 'A' + str(count_for_print) + ':' + 'D' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'полковник                                   В. Кузнецов'
    cell_prop.font = font3
    cell_prop.alignment = alignment1

    wb.save(lol_save)

def table_focus_on_swat(list_of_dict,path):
    lol_save = path
    lol_sheet = 'Worksheet'
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = lol_sheet

    # Шрифты для списка студентов(самый верх) и шапки колонки (Жирный)
    font_zirn = Font(name='Times New Roman',
                 size=14,
                 bold=True,  # жирный
                 italic=False,  # курсив
                 vertAlign='baseline',  # по какому краю выравнивать
                 # (['subscript', 'baseline', 'superscript'])
                 underline='none',  # подчеркивание написанного
                 strike=False,  # зачеркнуть
                 color='FF000000')

    # шрифты для подавших заявление на ВК2..... (не жирный)
    font1 = Font(name='Times New Roman',
                 size=14,
                 bold=False,  # жирный
                 italic=False,  # курсив
                 vertAlign='baseline',  # по какому краю выравнивать
                 # (['subscript', 'baseline', 'superscript'])
                 underline='none',  # подчеркивание написанного
                 strike=False,  # зачеркнуть
                 color='FF000000')

    # выравнивание для шапки таблицы
    alignment_centre = Alignment(horizontal='center',  # выравнивание по горизонтали
                           vertical='center',  # выравнивание по вертикали
                           text_rotation=0,
                           wrap_text=False,  # перенос текста
                           shrink_to_fit=False,
                           indent=0)

    # выравнивание (ыравние по левый)
    alignment_left = Alignment(horizontal='left',  # выравнивание по горизонтали
                           vertical='center',  # выравнивание по вертикали
                           text_rotation=0,
                           wrap_text=True,  # перенос текста
                           shrink_to_fit=False,
                           indent=0)

    # границы заливки (заливка таблицы)
    border1 = Border(left=Side(border_style='thin',
                               color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000'))

    # Ширина колонок
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 45
    ws1.column_dimensions['C'].width = 13
    ws1.column_dimensions['D'].width = 80

    # Заполнение клонок
    ws1.merge_cells('A1:D1')
    cell_prop = ws1['A1']
    cell_prop.value = u'СПИСОК СТУДЕНТОВ'
    cell_prop.font = font_zirn
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('A2:D2')
    cell_prop = ws1['A2']
    cell_prop.value = u'проходящих военное обучение по программам подготовки офицеров запаса на ВК №2'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('A3:D3')
    cell_prop = ws1['A3']
    cell_prop.value = u'подлежащим направлению на учебные сборы в 2016 году'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    cell_prop = ws1['A5']
    cell_prop.value = u'№'
    cell_prop.font = font_zirn
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['B5']
    cell_prop.value = u'ФИО'
    cell_prop.font = font_zirn
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['C5']
    cell_prop.value = u'Группа'
    cell_prop.font = font_zirn
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['D5']
    cell_prop.value = u'Военкомат'
    cell_prop.font = font_zirn
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    # получение списка военкоматов и сортировка его
    s_voenkom = []
    count = 0
    n = len(list_of_dict)
    while count < n:
        if not (list_of_dict[count]['title'] in s_voenkom):
            s_voenkom.append(list_of_dict[count]['title'])
        count += 1
    s_voenkom.sort()

    # получение списка групп и сортировка его
    list_of_group = []
    count = 0

    list_full_group = []
    while count < n:
        my_gr = ''
        my_gr = list_of_dict[count]['faculty'] + list_of_dict[count]['department'] +\
        '-' + list_of_dict[count]['group']
        list_full_group.append(my_gr)
        count += 1

    count = 0
    while count < n:
        if not (list_full_group[count] in list_of_group):
            list_of_group.append(list_full_group[count])
        count += 1
    list_of_group.sort()

    # вывод таблицы
    count_for_print = 6
    for voenk in s_voenkom:
        for group in list_of_group:
            count = 0
            first_last_three_names = []
            while count < n:
                if (voenk == list_of_dict[count]['title']) and (group == list_full_group[count]):
                    sleeve = list_of_dict[count]['last_name'] + ' ' + list_of_dict[count]['first_name'] + ' ' +\
                             list_of_dict[count]['middle_name']
                    first_last_three_names.append(sleeve)
                count += 1
            first_last_three_names.sort()
            for name in first_last_three_names:
                cell_cell = 'A' + str(count_for_print)
                cell_prop = ws1[cell_cell]
                cell_prop.value = count_for_print - 5
                cell_prop.font = font1
                cell_prop.alignment = alignment_centre
                cell_prop.border = border1

                cell_cell = 'B' + str(count_for_print)
                cell_prop = ws1[cell_cell]
                cell_prop.value = name
                cell_prop.font = font1
                cell_prop.alignment = alignment_left
                cell_prop.border = border1

                cell_cell = 'C' + str(count_for_print)
                cell_prop = ws1[cell_cell]
                cell_prop.value = group
                cell_prop.font = font1
                cell_prop.alignment = alignment_left
                cell_prop.border = border1

                cell_cell = 'D' + str(count_for_print)
                cell_prop = ws1[cell_cell]
                cell_prop.value = voenk
                cell_prop.font = font1
                cell_prop.alignment = alignment_left
                cell_prop.border = border1

                count_for_print += 1

    # Конец
    count_for_print += 1
    cell_cell = 'A' + str(count_for_print) + ':' + 'D' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'Начальник военной кафедры №2'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    count_for_print += 1
    cell_cell = 'A' + str(count_for_print) + ':' + 'D' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'полковник                                   В. Кузнецов'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    wb.save(lol_save)


def big_list(list_of_dict, path):
    lol_save = path
    lol_sheet = 'Worksheet'
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = lol_sheet

    # шрифты для подавших заявление на ВК2..... (не жирный)
    font1 = Font(name='Times New Roman',
                 size=12,
                 bold=False,  # жирный
                 italic=False,  # курсив
                 vertAlign='baseline',  # по какому краю выравнивать
                 # (['subscript', 'baseline', 'superscript'])
                 underline='none',  # подчеркивание написанного
                 strike=False,  # зачеркнуть
                 color='FF000000')

    # выравнивание для шапки таблицы
    alignment_centre = Alignment(horizontal='center',  # выравнивание по горизонтали
                                 vertical='center',  # выравнивание по вертикали
                                 text_rotation=0,
                                 wrap_text=True,  # перенос текста
                                 shrink_to_fit=False,
                                 indent=0)

    # выравнивание (ыравние по левый)
    alignment_left = Alignment(horizontal='left',  # выравнивание по горизонтали
                               vertical='center',  # выравнивание по вертикали
                               text_rotation=0,
                               wrap_text=True,  # перенос текста
                               shrink_to_fit=False,
                               indent=0)

    # границы заливки (заливка таблицы)
    border1 = Border(left=Side(border_style='thin',
                               color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000'))

    # Ширина колонок
    _cell = ws1.cell(row=1, column=1)
    _cell = ws1.cell(row=1, column=2)
    _cell = ws1.cell(row=1, column=3)
    _cell = ws1.cell(row=1, column=4)
    _cell = ws1.cell(row=1, column=5)
    _cell = ws1.cell(row=1, column=6)
    _cell = ws1.cell(row=1, column=7)
    _cell = ws1.cell(row=1, column=8)
    _cell = ws1.cell(row=1, column=9)
    _cell = ws1.cell(row=1, column=10)
    _cell = ws1.cell(row=1, column=11)

    ws1.column_dimensions['A'].width = 3
    ws1.column_dimensions['B'].width = 40
    ws1.column_dimensions['C'].width = 11
    ws1.column_dimensions['D'].width = 11
    ws1.column_dimensions['E'].width = 75
    ws1.column_dimensions['F'].width = 11
    ws1.column_dimensions['G'].width = 60
    ws1.column_dimensions['H'].width = 31
    ws1.column_dimensions['I'].width = 16
    ws1.column_dimensions['J'].width = 105
    ws1.column_dimensions['K'].width = 10

    # Заполнение шапки таблицы
    cell_prop = ws1['A1']
    cell_prop.value = u'№'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['B1']
    cell_prop.value = u'Фамилия Имя Отчество'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['C1']
    cell_prop.value = u'Учебная группа'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['D1']
    cell_prop.value = u'Дата рождения'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['E1']
    cell_prop.value = u'Военкомат'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['F1']
    cell_prop.value = u'Cпециальность'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['G1']
    cell_prop.value = u'Домашний адрес (по паспорту)'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['H1']
    cell_prop.value = u'Место рождения'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['I1']
    cell_prop.value = u'Серия и номер паспорта'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['J1']
    cell_prop.value = u'Кем и когда выдан паспорт'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['K1']
    cell_prop.value = u'Степень годности'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    # получение списка групп и сортировка его
    s = []
    count = 0
    n = len(list_of_dict)

    list_full_group = []
    while count < n:
        my_gr = ''
        my_gr = list_of_dict[count]['faculty'] + list_of_dict[count]['department'] +\
        '-' + list_of_dict[count]['group']
        list_full_group.append(my_gr)
        count += 1

    count = 0

    while count < n:
        if not (list_full_group[count] in s):
            s.append(list_full_group[count])
        count += 1
    s.sort()

    # вывод таблицы
    count_for_print = 2
    for group in s:
        count = 0
        first_last_three_names = []
        while count < n:
            if group == list_full_group[count]:
                sleeve = list_of_dict[count]['last_name'] + ' ' + list_of_dict[count]['first_name'] + ' ' +\
                         list_of_dict[count]['middle_name']
                first_last_three_names.append(sleeve)
            count += 1
        first_last_three_names.sort()
        for name in first_last_three_names:
            count_from_names = find_counter_in_list(name, list_of_dict)
            cell_cell = 'A' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = count_for_print - 1
            cell_prop.font = font1
            cell_prop.alignment = alignment_centre
            cell_prop.border = border1

            cell_cell = 'B' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = name
            cell_prop.font = font1
            cell_prop.alignment = alignment_left
            cell_prop.border = border1

            cell_cell = 'C' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = group
            cell_prop.font = font1
            cell_prop.alignment = alignment_left
            cell_prop.border = border1

            cell_cell = 'D' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = str(list_of_dict[count_from_names]['birth_date'])
            cell_prop.font = font1
            cell_prop.alignment = alignment_left
            cell_prop.border = border1

            cell_cell = 'E' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = list_of_dict[count_from_names]['title']
            cell_prop.font = font1
            cell_prop.alignment = alignment_left
            cell_prop.border = border1

            cell_cell = 'F' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = list_of_dict[count_from_names]['OKSO']
            cell_prop.font = font1
            cell_prop.alignment = alignment_left
            cell_prop.border = border1

            cell_cell = 'G' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = list_of_dict[count_from_names]['address']
            cell_prop.font = font1
            cell_prop.alignment = alignment_left
            cell_prop.border = border1

            cell_cell = 'H' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = list_of_dict[count_from_names]['birth_place']
            cell_prop.font = font1
            cell_prop.alignment = alignment_left
            cell_prop.border = border1

            cell_cell = 'I' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = list_of_dict[count_from_names]['document_number']
            cell_prop.font = font1
            cell_prop.alignment = alignment_left
            cell_prop.border = border1

            cell_cell = 'J' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = list_of_dict[count_from_names]['issued_by']
            cell_prop.font = font1
            cell_prop.alignment = alignment_left
            cell_prop.border = border1

            cell_cell = 'K' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = list_of_dict[count_from_names]['degree_of_fitness']
            cell_prop.font = font1
            cell_prop.alignment = alignment_left
            cell_prop.border = border1

            count_for_print += 1

    wb.save(lol_save)

def list_for_vus(list_of_dict,path):
    lol_save = path
    lol_sheet = 'Worksheet'
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = lol_sheet

    # шрифты для подавших заявление на ВК2..... (не жирный)
    font1 = Font(name='Times New Roman',
                 size=14,
                 bold=False,  # жирный
                 italic=False,  # курсив
                 vertAlign='baseline',  # по какому краю выравнивать
                 # (['subscript', 'baseline', 'superscript'])
                 underline='none',  # подчеркивание написанного
                 strike=False,  # зачеркнуть
                 color='FF000000')

    font_zirn = Font(name='Times New Roman',
                 size=14,
                 bold=True,  # жирный
                 italic=False,  # курсив
                 vertAlign='baseline',  # по какому краю выравнивать
                 # (['subscript', 'baseline', 'superscript'])
                 underline='none',  # подчеркивание написанного
                 strike=False,  # зачеркнуть
                 color='FF000000')

    # выравнивание для шапки таблицы
    alignment_centre = Alignment(horizontal='center',  # выравнивание по горизонтали
                                 vertical='center',  # выравнивание по вертикали
                                 text_rotation=0,
                                 wrap_text=False,  # перенос текста
                                 shrink_to_fit=False,
                                 indent=0)

    # выравнивание (ыравние по левый)
    alignment_left = Alignment(horizontal='left',  # выравнивание по горизонтали
                               vertical='center',  # выравнивание по вертикали
                               text_rotation=0,
                               wrap_text=False,  # перенос текста
                               shrink_to_fit=False,
                               indent=0)

    # границы заливки (заливка таблицы)
    border2 = Border(left=Side(border_style='thin',
                               color='FF000000'),
                     right=Side(border_style=None,
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000'))

    # границы заливки (заливка таблицы)
    border3 = Border(left=Side(border_style=None,
                               color='FF000000'),
                     right=Side(border_style=None,
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000'))

    # границы заливки (заливка таблицы)
    border4 = Border(left=Side(border_style=None,
                               color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000'))

    # границы заливки (заливка таблицы)
    border1 = Border(left=Side(border_style='thin',
                               color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000'))

    # Ширина колонок
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 62
    ws1.column_dimensions['C'].width = 16


    list_zvani_header = [u'проходящих военное обучение по программам подготовки солдат запаса',
                         u'проходящих военное обучение по программам подготовки сержантов запаса',
                         u'проходящих военное обучение по программам подготовки офицеров запаса']

    list_zvani = ['soldier', 'sergeant', 'officer']

    count_for_print_cell = 1
    zvanie = 0

    size_page = 37
    count_cell_page = 0
    while zvanie < 3:


        list_of_zvanie = []
        n = len(list_of_dict)
        count = 0
        while count < n:
            if list_of_dict[count]['grade'] == list_zvani[zvanie]:
                list_of_zvanie.append(list_of_dict[count])
            count += 1

        # Описание верхушки
        cells_position = 'A' + str(count_for_print_cell) + ':' + 'C' + str(count_for_print_cell)
        ws1.merge_cells(cells_position)
        cells_position = 'A' + str(count_for_print_cell)
        cell_prop = ws1[cells_position]
        cell_prop.value = u'СПИСОК СТУДЕНТОВ,'
        cell_prop.alignment = alignment_centre
        cell_prop.font = font_zirn

        count_for_print_cell += 1
        count_cell_page += 1


        if count_cell_page > size_page:
            count_for_print_cell = print_table_header_vus(ws1, count_for_print_cell)
            count_cell_page = 1

        cells_position = 'A' + str(count_for_print_cell) + ':' + 'C' + str(count_for_print_cell)
        ws1.merge_cells(cells_position)
        cells_position = 'A' + str(count_for_print_cell)
        cell_prop = ws1[cells_position]
        cell_prop.value = list_zvani_header[zvanie]
        cell_prop.alignment = alignment_centre
        cell_prop.font = font1

        count_for_print_cell += 1
        count_cell_page += 1
        if count_cell_page > size_page:
            count_for_print_cell = print_table_header_vus(ws1, count_for_print_cell)
            count_cell_page = 1

        cells_position = 'A' + str(count_for_print_cell) + ':' + 'C' + str(count_for_print_cell)
        ws1.merge_cells(cells_position)
        cells_position = 'A' + str(count_for_print_cell)
        cell_prop = ws1[cells_position]
        cell_prop.value = u'на ВК №2 в 2015-2016 учебном году'
        cell_prop.alignment = alignment_centre
        cell_prop.font = font1

        count_for_print_cell += 1
        count_cell_page += 1
        if count_cell_page > size_page:
            count_for_print_cell = print_table_header_vus(ws1, count_for_print_cell)
            count_cell_page = 1

        cells_position = 'A' + str(count_for_print_cell) + ':' + 'C' + str(count_for_print_cell)
        ws1.merge_cells(cells_position)
        cells_position = 'A' + str(count_for_print_cell)
        cell_prop = ws1[cells_position]
        cell_prop.value = ' '
        cell_prop.font = font1

        count_for_print_cell += 1
        count_cell_page += 1
        if count_cell_page > size_page:
            count_for_print_cell = print_table_header_vus(ws1, count_for_print_cell)
            count_cell_page = 1

        count_for_print_cell = print_table_header_vus(ws1, count_for_print_cell)
        count_cell_page += 1


        list_year = [u'ПЕРВЫЙ ГОД ОБУЧЕНИЯ', u'ВТОРОЙ ГОД ОБУЧЕНИЯ', u'ТРЕТИЙ ГОД ОБУЧЕНИЯ']

        year = 1
        while year <= 3:
            # Получение общей таблицы
            cells_position = 'A' + str(count_for_print_cell) + ':' + 'C' + str(count_for_print_cell)
            ws1.merge_cells(cells_position)
            cells_position = 'A' + str(count_for_print_cell)
            cell_prop = ws1[cells_position]
            cell_prop.border = border2
            cells_position = 'B' + str(count_for_print_cell)
            cell_prop = ws1[cells_position]
            cell_prop.border = border3
            cells_position = 'C' + str(count_for_print_cell)
            cell_prop = ws1[cells_position]
            cell_prop.border = border4

            cells_position = 'A' + str(count_for_print_cell)
            cell_prop = ws1[cells_position]
            cell_prop.value = list_year[year - 1]
            cell_prop.alignment = alignment_centre
            cell_prop.font = font_zirn

            count_for_print_cell += 1
            count_cell_page += 1
            if count_cell_page > size_page:
                count_for_print_cell = print_table_header_vus(ws1, count_for_print_cell)
                count_cell_page = 1

            # Заполнение таблицы людьми

            list_of_dict_year = []
            list_vus = []
            list_of_group = []
            n = len(list_of_zvanie)
            count = 0

            list_full_group = []
            while count < n:
                my_gr = ''
                my_gr = list_of_dict[count]['faculty'] + list_of_dict[count]['department'] +\
                '-' + list_of_dict[count]['group']
                list_full_group.append(my_gr)
                count += 1

            count = 0

            while count < n:
                if list_of_zvanie[count]['year'] == year:
                    list_of_dict_year.append(list_of_zvanie[count])
                    if not (list_of_zvanie[count]['VUS'] in list_vus):
                        list_vus.append(list_of_zvanie[count]['VUS'])
                    if not (list_full_group[count]  in list_of_group):
                        list_of_group.append(list_full_group[count])
                count += 1

            list_vus.sort()
            list_of_group.sort()
            n = len(list_of_dict_year)
            for vus in list_vus:

                count_for_print_name = 1

                cells_position = 'A' + str(count_for_print_cell) + ':' + 'C' + str(count_for_print_cell)
                ws1.merge_cells(cells_position)
                cells_position = 'A' + str(count_for_print_cell)
                cell_prop = ws1[cells_position]
                cell_prop.border = border2
                cells_position = 'B' + str(count_for_print_cell)
                cell_prop = ws1[cells_position]
                cell_prop.border = border3
                cells_position = 'C' + str(count_for_print_cell)
                cell_prop = ws1[cells_position]
                cell_prop.border = border4

                cells_position = 'A' + str(count_for_print_cell)
                cell_prop = ws1[cells_position]
                cell_prop.value = u'ВУС ' + vus.upper()
                cell_prop.alignment = alignment_centre
                cell_prop.font = font_zirn

                count_for_print_cell += 1
                count_cell_page += 1
                if count_cell_page > size_page:
                    count_for_print_cell = print_table_header_vus(ws1, count_for_print_cell)
                    count_cell_page = 1

                for group in list_of_group:
                    count = 0
                    first_last_three_names = []
                    while count < n:
                        if (vus == list_of_dict_year[count]['VUS']) and (group == list_full_group[count]):
                            sleeve = list_of_dict_year[count]['last_name'] + ' ' + list_of_dict_year[count]['first_name'] + ' '+\
                                     list_of_dict_year[count]['middle_name']
                            first_last_three_names.append(sleeve)
                        count += 1
                    first_last_three_names.sort()
                    for name in first_last_three_names:
                        cell_cell = 'A' + str(count_for_print_cell)
                        cell_prop = ws1[cell_cell]
                        cell_prop.value = str(count_for_print_name) + '.'
                        cell_prop.font = font1
                        cell_prop.alignment = alignment_centre
                        cell_prop.border = border1
                        count_for_print_name += 1

                        cell_cell = 'B' + str(count_for_print_cell)
                        cell_prop = ws1[cell_cell]
                        cell_prop.value = name
                        cell_prop.font = font1
                        cell_prop.alignment = alignment_left
                        cell_prop.border = border1

                        cell_cell = 'C' + str(count_for_print_cell)
                        cell_prop = ws1[cell_cell]
                        cell_prop.value = group
                        cell_prop.font = font1
                        cell_prop.alignment = alignment_centre
                        cell_prop.border = border1

                        count_for_print_cell += 1
                        count_cell_page += 1
                        if count_cell_page > size_page:
                            count_for_print_cell = print_table_header_vus(ws1, count_for_print_cell)
                            count_cell_page = 1
            year += 1

        # Конец
        cells_position = 'A' + str(count_for_print_cell) + ':' + 'C' + str(count_for_print_cell)
        ws1.merge_cells(cells_position)
        cells_position = 'A' + str(count_for_print_cell)
        cell_prop = ws1[cells_position]
        cell_prop.value = ' '
        cell_prop.font = font1

        count_for_print_cell += 1
        count_cell_page += 1
        if count_cell_page > size_page:
            cells_position = 'A' + str(count_for_print_cell) + ':' + 'C' + str(count_for_print_cell)
            ws1.merge_cells(cells_position)
            cells_position = 'A' + str(count_for_print_cell)
            cell_prop = ws1[cells_position]
            cell_prop.value = ' '
            cell_prop.font = font1
            count_cell_page = 1
            count_for_print_cell += 1

        cell_cell = 'A' + str(count_for_print_cell) + ':' + 'C' + str(count_for_print_cell)
        ws1.merge_cells(cell_cell)
        cell_cell = 'A' + str(count_for_print_cell)
        cell_prop = ws1[cell_cell]
        cell_prop.value = u'Начальник военной кафедры №2'
        cell_prop.font = font1
        cell_prop.alignment = alignment_centre

        count_for_print_cell += 1
        count_cell_page += 1
        if count_cell_page > size_page:
            cells_position = 'A' + str(count_for_print_cell) + ':' + 'C' + str(count_for_print_cell)
            ws1.merge_cells(cells_position)
            cells_position = 'A' + str(count_for_print_cell)
            cell_prop = ws1[cells_position]
            cell_prop.value = ' '
            cell_prop.font = font1
            count_cell_page = 1
            count_for_print_cell += 1

        cell_cell = 'A' + str(count_for_print_cell) + ':' + 'C' + str(count_for_print_cell)
        ws1.merge_cells(cell_cell)
        cell_cell = 'A' + str(count_for_print_cell)
        cell_prop = ws1[cell_cell]
        cell_prop.value = u'полковник                                   В. Кузнецов'
        cell_prop.font = font1
        cell_prop.alignment = alignment_centre
        count_for_print_cell += 1
        count_cell_page += 1


        while count_cell_page < size_page + 1:
            cells_position = 'A' + str(count_for_print_cell) + ':' + 'C' + str(count_for_print_cell)
            ws1.merge_cells(cells_position)
            cells_position = 'A' + str(count_for_print_cell)
            cell_prop = ws1[cells_position]
            cell_prop.value = ' '
            cell_prop.font = font1
            count_cell_page += 1
            count_for_print_cell += 1

        count_cell_page = 0
        zvanie += 1




    wb.save(lol_save)

def print_table_header_military_training(ws, cell_pos):
    # шрифты для подавших заявление на ВК2..... (не жирный)
    font1 = Font(name='Times New Roman',
                 size=12,
                 bold=False,  # жирный
                 italic=False,  # курсив
                 vertAlign='baseline',  # по какому краю выравнивать
                 # (['subscript', 'baseline', 'superscript'])
                 underline='none',  # подчеркивание написанного
                 strike=False,  # зачеркнуть
                 color='FF000000')

    font_zirn = Font(name='Times New Roman',
                     size=14,
                     bold=True,  # жирный
                     italic=False,  # курсив
                     vertAlign='baseline',  # по какому краю выравнивать
                     # (['subscript', 'baseline', 'superscript'])
                     underline='none',  # подчеркивание написанного
                     strike=False,  # зачеркнуть
                     color='FF000000')

    # выравнивание для шапки таблицы
    alignment_centre = Alignment(horizontal='center',  # выравнивание по горизонтали
                                 vertical='center',  # выравнивание по вертикали
                                 text_rotation=0,
                                 wrap_text=False,  # перенос текста
                                 shrink_to_fit=False,
                                 indent=0)

    # выравнивание для шапки таблицы
    alignment_centre_wrap_text = Alignment(horizontal='center',  # выравнивание по горизонтали
                                           vertical='center',  # выравнивание по вертикали
                                           text_rotation=0,
                                           wrap_text=True,  # перенос текста
                                           shrink_to_fit=False,
                                           indent=0)

    # выравнивание (ыравние по левый)
    alignment_left = Alignment(horizontal='left',  # выравнивание по горизонтали
                               vertical='center',  # выравнивание по вертикали
                               text_rotation=0,
                               wrap_text=True,  # перенос текста
                               shrink_to_fit=False,
                               indent=0)

    # выравнивание (ыравние по правый)
    alignment_right = Alignment(horizontal='right',  # выравнивание по горизонтали
                                vertical='center',  # выравнивание по вертикали
                                text_rotation=0,
                                wrap_text=False,  # перенос текста
                                shrink_to_fit=False,
                                indent=0)

    # Выровнять по ширине (Провить кое что)
    alignment_width = Alignment(horizontal='fill',  # выравнивание по горизонтали
                                vertical='center',  # выравнивание по вертикали
                                text_rotation=0,
                                wrap_text=False,  # перенос текста
                                shrink_to_fit=False,
                                indent=0)

    # границы заливки (заливка таблицы)
    border2 = Border(left=Side(border_style='thin',
                               color='FF000000'),
                     right=Side(border_style=None,
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000'))

    # границы заливки (заливка таблицы)
    border3 = Border(left=Side(border_style=None,
                               color='FF000000'),
                     right=Side(border_style=None,
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000'))

    # границы заливки (заливка таблицы)
    border4 = Border(left=Side(border_style=None,
                               color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000'))

    # границы заливки (заливка таблицы)
    border1 = Border(left=Side(border_style='thin',
                               color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000'))

    # Формируем название полей в таблице

    cell_cell = 'A' + str(cell_pos)
    cell_prop = ws[cell_cell]
    cell_prop.value = u'№ п/п'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre_wrap_text
    cell_prop.border = border1

    cell_cell = 'B' + str(cell_pos)
    cell_prop = ws[cell_cell]
    cell_prop.value = u'Фамилия, имя и отчество'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre_wrap_text
    cell_prop.border = border1

    cell_cell = 'C' + str(cell_pos)
    cell_prop = ws[cell_cell]
    cell_prop.value = u'Факультет'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre_wrap_text
    cell_prop.border = border1

    cell_cell = 'D' + str(cell_pos)
    cell_prop = ws[cell_cell]
    cell_prop.value = u'Учебная группа'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre_wrap_text
    cell_prop.border = border1

    cell_cell = 'E' + str(cell_pos)
    cell_prop = ws[cell_cell]
    cell_prop.value = u'В каком военном комиссариате состоит на воинском учете'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre_wrap_text
    cell_prop.border = border1

    cell_cell = 'F' + str(cell_pos)
    cell_prop = ws[cell_cell]
    cell_prop.value = u'Примеч.'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre_wrap_text
    cell_prop.border = border1

    cell_pos += 1


def table_military_training(list_of_dict,path):
    lol_save = path
    lol_sheet = 'Worksheet'
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = lol_sheet

    # шрифты для подавших заявление на ВК2..... (не жирный)
    font1 = Font(name='Times New Roman',
                 size=12,
                 bold=False,  # жирный
                 italic=False,  # курсив
                 vertAlign='baseline',  # по какому краю выравнивать
                 # (['subscript', 'baseline', 'superscript'])
                 underline='none',  # подчеркивание написанного
                 strike=False,  # зачеркнуть
                 color='FF000000')


    # выравнивание для шапки таблицы
    alignment_centre = Alignment(horizontal='center',  # выравнивание по горизонтали
                                 vertical='center',  # выравнивание по вертикали
                                 text_rotation=0,
                                 wrap_text=False,  # перенос текста
                                 shrink_to_fit=False,
                                 indent=0)

    # выравнивание для шапки таблицы
    alignment_centre_wrap_text = Alignment(horizontal='center',  # выравнивание по горизонтали
                                           vertical='center',  # выравнивание по вертикали
                                           text_rotation=0,
                                           wrap_text=True,  # перенос текста
                                           shrink_to_fit=False,
                                           indent=0)

    # выравнивание (ыравние по левый)
    alignment_left = Alignment(horizontal='left',  # выравнивание по горизонтали
                               vertical='center',  # выравнивание по вертикали
                               text_rotation=0,
                               wrap_text=True,  # перенос текста
                               shrink_to_fit=False,
                               indent=0)

    # выравнивание (ыравние по правый)
    alignment_right = Alignment(horizontal='right',  # выравнивание по горизонтали
                               vertical='center',  # выравнивание по вертикали
                               text_rotation=0,
                               wrap_text=False,  # перенос текста
                               shrink_to_fit=False,
                               indent=0)

    # Выровнять по ширине (Провить кое что)
    alignment_width = Alignment(horizontal='justify',  # выравнивание по горизонтали
                                vertical='center',  # выравнивание по вертикали
                                text_rotation=0,
                                wrap_text=False,  # перенос текста
                                shrink_to_fit=False,
                                indent=0)

    # границы заливки (заливка таблицы)
    border1 = Border(left=Side(border_style='thin',
                               color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000'))

    # Ширина колонок
    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 11
    ws1.column_dimensions['D'].width = 11
    ws1.column_dimensions['E'].width = 35
    ws1.column_dimensions['F'].width = 9

    ws1.merge_cells('D1:F1')
    cell_prop = ws1['D1']
    cell_prop.value = u'УТВЕРЖДАЮ'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('D2:F2')
    cell_prop = ws1['D2']
    cell_prop.value = u'Ректор Московского Государственного'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('D3:F3')
    cell_prop = ws1['D3']
    cell_prop.value = u'технического университета им. Н.Э.Бауман'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('D4:F4')
    cell_prop = ws1['D4']
    cell_prop.value = u''
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('D5:F5')
    cell_prop = ws1['D5']
    cell_prop.value = u'А.А. Александров'
    cell_prop.font = font1
    cell_prop.alignment = alignment_right

    ws1.merge_cells('D6:F6')
    cell_prop = ws1['D6']
    cell_prop.value = u''
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('A7:F7')
    cell_prop = ws1['A7']
    cell_prop.value = u'М.П.'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('A8:F8')
    cell_prop = ws1['A8']
    cell_prop.value = u''
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('A9:F9')
    cell_prop = ws1['A9']
    cell_prop.value = u''
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('A10:F10')
    cell_prop = ws1['A10']
    cell_prop.value = u''
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('A11:F11')
    cell_prop = ws1['A11']
    cell_prop.value = u'СПИСОК'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('A12:F12')
    cell_prop = ws1['A12']
    left_string = u'граждан из числа студентов, проходящих военную подготовку по '
    right_string = u' на военной'
    temp_string_to_vus = left_string + list_of_dict[0]['VUS'] + right_string
    cell_prop.value = temp_string_to_vus
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('A13:F13')
    cell_prop = ws1['A13']
    cell_prop.value = u'кафедре №2, направляемых на учебные сборы с 01.07.2015 г. по 03.08.2015 г.'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('A14:F14')
    cell_prop = ws1['A14']
    cell_prop.value = u'в войсковую часть 03523 (пгт. Софрино-1, Московская обл.)'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('A15:F15')
    cell_prop = ws1['A15']
    cell_prop.value = u''
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    # Формируем название полей в таблице
    cell_prop = ws1['A16']
    cell_prop.value = u'№ п/п'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre_wrap_text
    cell_prop.border = border1

    cell_prop = ws1['B16']
    cell_prop.value = u'Фамилия, имя и отчество'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre_wrap_text
    cell_prop.border = border1

    cell_prop = ws1['C16']
    cell_prop.value = u'Факультет'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre_wrap_text
    cell_prop.border = border1

    cell_prop = ws1['D16']
    cell_prop.value = u'Учебная группа'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre_wrap_text
    cell_prop.border = border1

    cell_prop = ws1['E16']
    cell_prop.value = u'В каком военном комиссариате состоит на воинском учете'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre_wrap_text
    cell_prop.border = border1

    cell_prop = ws1['F16']
    cell_prop.value = u'Примеч.'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre_wrap_text
    cell_prop.border = border1

    # Вывод для таблицы
    list_full_group = []

    s = []
    n = len(list_of_dict)
    count = 0
    while count < n:
        my_gr = list_of_dict[count]['faculty'] + list_of_dict[count]['department'] + \
                '-' + list_of_dict[count]['group']
        list_full_group.append(my_gr)
        count += 1

    count = 0
    while count < n:
        if not (list_full_group[count] in s):
            s.append(list_full_group[count])
        count += 1
    s.sort()

    count_header_print = 17
    count_for_print = 17
    for group in s:
        count = 0
        first_last_three_names = []
        while count < n:
            if group == list_full_group[count]:
                sleeve = list_of_dict[count]['last_name'] + ' ' + list_of_dict[count]['first_name'] + ' ' +\
                         list_of_dict[count]['middle_name']
                first_last_three_names.append(sleeve)
            count += 1
        first_last_three_names.sort()
        for name in first_last_three_names:
            count_from_names = find_counter_in_list(name, list_of_dict)
            cell_cell = 'A' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = str(count_for_print - 16) + '.'
            cell_prop.font = font1
            cell_prop.alignment = alignment_centre
            cell_prop.border = border1

            cell_cell = 'B' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = name
            cell_prop.font = font1
            cell_prop.alignment = alignment_left
            cell_prop.border = border1

            cell_cell = 'C' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = list_of_dict[count_from_names]['faculty']
            cell_prop.font = font1
            cell_prop.alignment = alignment_centre
            cell_prop.border = border1

            cell_cell = 'D' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = group
            cell_prop.font = font1
            cell_prop.alignment = alignment_centre
            cell_prop.border = border1

            cell_cell = 'E' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = list_of_dict[count_from_names]['title']
            cell_prop.font = font1
            cell_prop.alignment = alignment_left
            cell_prop.border = border1

            cell_cell = 'F' + str(count_for_print)
            cell_prop = ws1[cell_cell]
            cell_prop.value = ''
            cell_prop.font = font1
            cell_prop.alignment = alignment_centre
            cell_prop.border = border1

            count_for_print += 1
            # count_header_print += 1
            # if count_header_print > 40:
            #     count_header_print = 1
            #     print_table_header_military_training(ws1, count_for_print)
            #     count_for_print += 1


    # Вывод конец

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u''
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    temp_string = u'    Основание: Постановление Правительства РФ от 06 марта 2008 г. № 152,'
    cell_prop.value = temp_string
    cell_prop.font = font1
    cell_prop.alignment = alignment_width
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    temp_string = u'совместный приказ Минобороны и Минобрнауки РФ от 10 июня 2009 г. № 666/249,'
    cell_prop.value = temp_string
    cell_prop.font = font1
    cell_prop.alignment = alignment_width
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    temp_string = u'приказ Командующего войсками Западного военного округа от 28 ноября 2014 г.'
    cell_prop.value = temp_string
    cell_prop.font = font1
    cell_prop.alignment = alignment_width
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    temp_string = u'№ 683, сводный план проведения учебных сборов в 2015 г. Войск ВКО.'
    cell_prop.value = temp_string
    cell_prop.font = font1
    cell_prop.alignment = alignment_width
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u''
    cell_prop.font = font1
    cell_prop.alignment = alignment_width
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'Директор Военного института '
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'полковник _________________Н. Максименко'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u''
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'    Продовольственно-путевыми деньгами не обеспечивались.'
    cell_prop.font = font1
    cell_prop.alignment = alignment_width
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'    Имеют право для зачисления на все установленные виды довольствия'
    cell_prop.font = font1
    cell_prop.alignment = alignment_width
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'с момента прибытия в воинскую часть.'
    cell_prop.font = font1
    cell_prop.alignment = alignment_width
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u''
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'Начальник отдела (объединенного) военного комиссариата  г. Москвы'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'по Измайловскому району ВАО г. Москвы '
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'_______________ В. Кудряшов'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u''
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'    М.П.'
    cell_prop.font = font1
    cell_prop.alignment = alignment_left
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u''
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    temp_string_to_vus = u'    Поименованные в списке граждане, за исключением ____________'
    cell_prop.value = temp_string_to_vus
    cell_prop.font = font1
    cell_prop.alignment = alignment_width
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    temp_string_to_vus = u'_________________________________________________________________'
    cell_prop.value = temp_string_to_vus
    cell_prop.font = font1
    cell_prop.alignment = alignment_width
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    temp_string_to_vus = u'_________________________________________________________________'
    cell_prop.value = temp_string_to_vus
    cell_prop.font = font1
    cell_prop.alignment = alignment_width
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    temp_string_to_vus = u'________________________ в период с 01.07.2015 г. по 03.08.2015 г'
    cell_prop.value = temp_string_to_vus
    cell_prop.font = font1
    cell_prop.alignment = alignment_width
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    temp_string_to_vus = u'прошли учебный сбор в войсковой части 03523 (пгт. Софрино-1) по'
    cell_prop.value = temp_string_to_vus
    cell_prop.font = font1
    cell_prop.alignment = alignment_width
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    temp_string_to_vus = u'военно-учетной специальности ' + list_of_dict[0]['VUS'] + '.'
    cell_prop.value = temp_string_to_vus
    cell_prop.font = font1
    cell_prop.alignment = alignment_width
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u''
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'Командир войсковой части 03523'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'подполковник _______________ А. Чумаков'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre
    count_for_print += 1

    cell_cell = 'A' + str(count_for_print) + ':' + 'F' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    temp_string_to_vus = u'     М.П.'
    cell_prop.value = temp_string_to_vus
    cell_prop.font = font1
    cell_prop.alignment = alignment_width
    count_for_print += 1

    wb.save(lol_save)


def table_any(list_of_dict,path):
    lol_save = path
    lol_sheet = 'Worksheet'
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = lol_sheet

    # Шрифты для списка студентов(самый верх) и шапки колонки (Жирный)
    font_zirn = Font(name='Times New Roman',
                 size=14,
                 bold=True,  # жирный
                 italic=False,  # курсив
                 vertAlign='baseline',  # по какому краю выравнивать
                 # (['subscript', 'baseline', 'superscript'])
                 underline='none',  # подчеркивание написанного
                 strike=False,  # зачеркнуть
                 color='FF000000')

    # шрифты для подавших заявление на ВК2..... (не жирный)
    font1 = Font(name='Times New Roman',
                 size=14,
                 bold=False,  # жирный
                 italic=False,  # курсив
                 vertAlign='baseline',  # по какому краю выравнивать
                 # (['subscript', 'baseline', 'superscript'])
                 underline='none',  # подчеркивание написанного
                 strike=False,  # зачеркнуть
                 color='FF000000')

    # выравнивание для шапки таблицы
    alignment_centre = Alignment(horizontal='center',  # выравнивание по горизонтали
                           vertical='center',  # выравнивание по вертикали
                           text_rotation=0,
                           wrap_text=False,  # перенос текста
                           shrink_to_fit=False,
                           indent=0)

    # выравнивание (ыравние по левый)
    alignment_left = Alignment(horizontal='left',  # выравнивание по горизонтали
                           vertical='center',  # выравнивание по вертикали
                           text_rotation=0,
                           wrap_text=True,  # перенос текста
                           shrink_to_fit=False,
                           indent=0)

    # границы заливки (заливка таблицы)
    border1 = Border(left=Side(border_style='thin',
                               color='FF000000'),
                     right=Side(border_style='thin',
                                color='FF000000'),
                     top=Side(border_style='thin',
                              color='FF000000'),
                     bottom=Side(border_style='thin',
                                 color='FF000000'),
                     diagonal=Side(border_style=None,
                                   color='FF000000'),
                     diagonal_direction=0,
                     outline=Side(border_style=None,
                                  color='FF000000'),
                     vertical=Side(border_style=None,
                                   color='FF000000'),
                     horizontal=Side(border_style=None,
                                     color='FF000000'))

    # Ширина колонок
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 45
    ws1.column_dimensions['C'].width = 13
    ws1.column_dimensions['D'].width = 80
    ws1.column_dimensions['E'].width = 80
    ws1.column_dimensions['F'].width = 80
    ws1.column_dimensions['G'].width = 80
    ws1.column_dimensions['H'].width = 80
    ws1.column_dimensions['I'].width = 80
    ws1.column_dimensions['J'].width = 80
    ws1.column_dimensions['K'].width = 80
    ws1.column_dimensions['L'].width = 80
    ws1.column_dimensions['M'].width = 80

    list_all_keys = ['last_name',
                     'first_name',
                     'middle_name',
                     'birth_date',
                     'faculty',
                     'department',
                     'group',

                     'grade',
                     'year',
                     'birth_place',
                     'OKSO_letters',
                     'title',
                     'degree_of_fitness',
                     'document_number',

                     'issued_by',
                     'VUS',
                     'address',
                     'OKSO'
                     ]

    list_all_keys_width = [15, 15, 15, 11, 4, 4, 3,
                           12, 10, 40, 40, 80, 20, 100,
                           38, 9, 70, 10]

    

    # Заполнение клонок
    ws1.merge_cells('A1:D1')
    cell_prop = ws1['A1']
    cell_prop.value = u'СПИСОК СТУДЕНТОВ'
    cell_prop.font = font_zirn
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('A2:D2')
    cell_prop = ws1['A2']
    cell_prop.value = u'проходящих военное обучение по программам подготовки офицеров запаса на ВК №2'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    ws1.merge_cells('A3:D3')
    cell_prop = ws1['A3']
    cell_prop.value = u'подлежащим направлению на учебные сборы в 2016 году'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    cell_prop = ws1['A5']
    cell_prop.value = u'№'
    cell_prop.font = font_zirn
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['B5']
    cell_prop.value = u'ФИО'
    cell_prop.font = font_zirn
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['C5']
    cell_prop.value = u'Группа'
    cell_prop.font = font_zirn
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    cell_prop = ws1['D5']
    cell_prop.value = u'Военкомат'
    cell_prop.font = font_zirn
    cell_prop.alignment = alignment_centre
    cell_prop.border = border1

    # получение списка военкоматов и сортировка его
    s_voenkom = []
    count = 0
    n = len(list_of_dict)
    while count < n:
        if not (list_of_dict[count]['title'] in s_voenkom):
            s_voenkom.append(list_of_dict[count]['title'])
        count += 1
    s_voenkom.sort()

    # получение списка групп и сортировка его
    list_of_group = []
    count = 0

    list_full_group = []
    while count < n:
        my_gr = ''
        my_gr = list_of_dict[count]['faculty'] + list_of_dict[count]['department'] +\
        '-' + list_of_dict[count]['group']
        list_full_group.append(my_gr)
        count += 1

    count = 0
    while count < n:
        if not (list_full_group[count] in list_of_group):
            list_of_group.append(list_full_group[count])
        count += 1
    list_of_group.sort()

    # вывод таблицы
    count_for_print = 6
    for voenk in s_voenkom:
        for group in list_of_group:
            count = 0
            first_last_three_names = []
            while count < n:
                if (voenk == list_of_dict[count]['title']) and (group == list_full_group[count]):
                    sleeve = list_of_dict[count]['last_name'] + ' ' + list_of_dict[count]['first_name'] + ' ' +\
                             list_of_dict[count]['middle_name']
                    first_last_three_names.append(sleeve)
                count += 1
            first_last_three_names.sort()
            for name in first_last_three_names:
                cell_cell = 'A' + str(count_for_print)
                cell_prop = ws1[cell_cell]
                cell_prop.value = count_for_print - 5
                cell_prop.font = font1
                cell_prop.alignment = alignment_centre
                cell_prop.border = border1

                cell_cell = 'B' + str(count_for_print)
                cell_prop = ws1[cell_cell]
                cell_prop.value = name
                cell_prop.font = font1
                cell_prop.alignment = alignment_left
                cell_prop.border = border1

                cell_cell = 'C' + str(count_for_print)
                cell_prop = ws1[cell_cell]
                cell_prop.value = group
                cell_prop.font = font1
                cell_prop.alignment = alignment_left
                cell_prop.border = border1

                cell_cell = 'D' + str(count_for_print)
                cell_prop = ws1[cell_cell]
                cell_prop.value = voenk
                cell_prop.font = font1
                cell_prop.alignment = alignment_left
                cell_prop.border = border1

                count_for_print += 1

    # Конец
    count_for_print += 1
    cell_cell = 'A' + str(count_for_print) + ':' + 'D' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'Начальник военной кафедры №2'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    count_for_print += 1
    cell_cell = 'A' + str(count_for_print) + ':' + 'D' + str(count_for_print)
    ws1.merge_cells(cell_cell)
    cell_cell = 'A' + str(count_for_print)
    cell_prop = ws1[cell_cell]
    cell_prop.value = u'полковник                                   В. Кузнецов'
    cell_prop.font = font1
    cell_prop.alignment = alignment_centre

    wb.save(lol_save)